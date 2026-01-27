from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


TEMPLATE_DEFAULT = (Path(__file__).resolve().parents[1] / "training" / "LM5148_LM25148_quickstart_calculator_A4.xlsm")


def load_payload(json_path: Path) -> dict:
    payload = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or "inputs" not in payload:
        raise ValueError("JSON must contain an 'inputs' object")
    return payload


def fill_quickstart(
    template_path: Path,
    payload: dict,
    out_path: Path,
    out_xlsx: Path | None = None,
    use_excel: bool = False,
) -> None:
    # Known input cells on sheet 'Design Regulator' (observed in the template):
    # E6 VIN(min), E7 VIN(nom), E8 VIN(max), E9 VOUT, E10 IOUT, E11 FSW (kHz)
    inputs = payload.get("inputs", {})

    vin_min = float(inputs.get("vinMin", inputs.get("vinNom")))
    vin_nom = float(inputs.get("vinNom"))
    vin_max = float(inputs.get("vinMax"))
    vout = float(inputs.get("vout"))
    iout = float(inputs.get("iout"))
    fsw_hz = float(inputs.get("fsw"))

    fsw_khz = fsw_hz / 1000.0

    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Optional: automate Excel via xlwings to preserve formatting/shapes/macros and force recalculation.
    # This requires Windows + installed Excel + a Python that supports pywin32 (typically 3.10/3.11).
    if use_excel:
        try:
            import xlwings as xw

            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            app.screen_updating = False
            try:
                book = app.books.open(str(template_path), update_links=False, read_only=False)
                try:
                    sh = book.sheets["Design Regulator"]
                    sh.range("E6").value = vin_min
                    sh.range("E7").value = vin_nom
                    sh.range("E8").value = vin_max
                    sh.range("E9").value = vout
                    sh.range("E10").value = iout
                    sh.range("E11").value = fsw_khz

                    # Force recalculation so dependent sheets update.
                    try:
                        book.app.api.CalculateFullRebuild()
                    except Exception:
                        book.app.calculate()

                    book.save(str(out_path))

                    if out_xlsx is not None:
                        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
                        # 51 = xlOpenXMLWorkbook (.xlsx)
                        book.api.SaveAs(str(out_xlsx), FileFormat=51)
                finally:
                    book.close()
            finally:
                app.quit()
            return
        except Exception as exc:
            print(f"Warning: Excel automation failed, falling back to openpyxl: {exc}")

    def apply_inputs(wb: openpyxl.Workbook) -> None:
        ws = wb["Design Regulator"]
        ws["E6"].value = vin_min
        ws["E7"].value = vin_nom
        ws["E8"].value = vin_max
        ws["E9"].value = vout
        ws["E10"].value = iout
        ws["E11"].value = fsw_khz

    wb_xlsm = openpyxl.load_workbook(template_path, keep_vba=True, data_only=False)
    apply_inputs(wb_xlsm)
    wb_xlsm.save(out_path)

    if out_xlsx is not None:
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb_xlsx = openpyxl.load_workbook(template_path, keep_vba=False, data_only=False)
        apply_inputs(wb_xlsx)
        wb_xlsx.save(out_xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Fill the TI LM5148/LM25148 Quickstart Calculator (.xlsm) with values from the webapp JSON export. "
            "Saves a new .xlsm copy."
        )
    )
    parser.add_argument("--json", type=str, required=True, help="Path to lm5148_design.json downloaded from the webapp")
    parser.add_argument(
        "--template",
        type=str,
        default=str(TEMPLATE_DEFAULT),
        help="Path to LM5148_LM25148_quickstart_calculator_A4.xlsm",
    )
    parser.add_argument(
        "--out",
        type=str,
        default=str(Path.cwd() / "LM5148_quickstart_filled.xlsm"),
        help="Output .xlsm path",
    )
    parser.add_argument(
        "--out-xlsx",
        type=str,
        default="",
        help="Optional output .xlsx path (macros removed)",
    )
    parser.add_argument(
        "--use-excel",
        action="store_true",
        help=(
            "Use Excel automation (xlwings) to preserve shapes/macros and force recalculation before saving. "
            "Recommended to run under Python 3.10/3.11 with pywin32 installed."
        ),
    )

    args = parser.parse_args()

    json_path = Path(args.json)
    template_path = Path(args.template)
    out_path = Path(args.out)
    out_xlsx = Path(args.out_xlsx) if args.out_xlsx else None

    if not json_path.exists():
        raise SystemExit(f"JSON not found: {json_path}")
    if not template_path.exists():
        raise SystemExit(f"Template not found: {template_path}")

    payload = load_payload(json_path)
    fill_quickstart(template_path, payload, out_path, out_xlsx=out_xlsx, use_excel=bool(args.use_excel))

    print(f"Wrote: {out_path}")
    if out_xlsx is not None:
        print(f"Wrote: {out_xlsx}")
    print("Note: Excel will recalculate formulas when opened.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
