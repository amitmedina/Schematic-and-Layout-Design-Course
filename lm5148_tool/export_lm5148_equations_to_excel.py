import argparse
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Tuple

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class EquationItem:
    eq_id: int
    page: int  # 1-based
    text: str
    context: str
    bbox: Tuple[float, float, float, float]  # (x0,y0,x1,y1) in PDF points
    image_path: str


_EQUATION_HINT_RE = re.compile(
    # Keep this intentionally strict to avoid classifying plain specs/prose as "equations".
    r"(=|≤|≥|≈|≠|/|∑|√|\^)"
)


def normalize_equation_text(text: str) -> str:
    t = " ".join(text.split())
    # Common normalization for minus and weird spacing
    t = t.replace("−", "-")
    return t.strip()


def looks_like_equation(text: str, *, min_len: int, require_equals: bool) -> bool:
    t = normalize_equation_text(text)
    if len(t) < min_len:
        return False

    has_math_hint = bool(_EQUATION_HINT_RE.search(t))
    if not has_math_hint:
        return False

    if require_equals and "=" not in t and "≤" not in t and "≥" not in t and "≈" not in t and "≠" not in t:
        # Most datasheet equations use explicit relation operators.
        return False

    # Filter obvious prose blocks
    words = t.split()
    if len(words) >= 16 and not any(ch in t for ch in ["=", "/", "∑", "√"]):
        return False

    # Require at least one letter variable-ish OR a V/I/R style token
    has_variable = bool(re.search(r"[A-Za-z]", t))
    has_number = bool(re.search(r"\d", t))

    # Many datasheet callouts are like "VOUT = 5 V" (letters + digits + '=')
    return (has_variable and has_math_hint) and (has_number or "=" in t)


def extract_text_blocks(page: fitz.Page) -> List[dict]:
    # Use "blocks" for speed; it's much faster than the nested dict/lines/spans structure.
    # Each entry: (x0, y0, x1, y1, "text", block_no, block_type)
    blocks = []
    for b in page.get_text("blocks"):
        if len(b) < 7:
            continue
        x0, y0, x1, y1, text, _block_no, block_type = b[:7]
        if block_type != 0:
            continue
        if not text or not str(text).strip():
            continue
        blocks.append(
            {
                "type": 0,
                "bbox": [float(x0), float(y0), float(x1), float(y1)],
                "text": str(text),
            }
        )
    return blocks


def block_text(block: dict) -> str:
    return str(block.get("text", ""))


def block_bbox(block: dict) -> Tuple[float, float, float, float]:
    bbox = block.get("bbox", [0, 0, 0, 0])
    return float(bbox[0]), float(bbox[1]), float(bbox[2]), float(bbox[3])


def sorted_blocks_by_position(blocks: List[dict]) -> List[dict]:
    def key(b: dict) -> Tuple[float, float]:
        x0, y0, *_ = block_bbox(b)
        return (y0, x0)

    return sorted(blocks, key=key)


def build_context(blocks_sorted: List[dict], idx: int) -> str:
    # Very short, helpful context: previous and next block snippets
    prev_txt = ""
    next_txt = ""

    if idx - 1 >= 0:
        prev_txt = normalize_equation_text(block_text(blocks_sorted[idx - 1]))
    if idx + 1 < len(blocks_sorted):
        next_txt = normalize_equation_text(block_text(blocks_sorted[idx + 1]))

    def clip(s: str) -> str:
        s = " ".join(s.split())
        if len(s) > 140:
            return s[:137] + "..."
        return s

    prev_txt = clip(prev_txt)
    next_txt = clip(next_txt)

    if prev_txt and next_txt:
        return f"{prev_txt} | {next_txt}"
    return prev_txt or next_txt or ""


def render_crop(page: fitz.Page, rect: fitz.Rect, *, zoom: float, out_path: Path) -> None:
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, clip=rect, alpha=False)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    pix.save(out_path.as_posix())


def dedupe(items: Iterable[Tuple[int, str, Tuple[float, float, float, float]]]) -> List[Tuple[int, str, Tuple[float, float, float, float]]]:
    seen = set()
    out = []
    for page_num, text, bbox in items:
        t = normalize_equation_text(text)
        # bbox rounding to reduce tiny differences
        bb = tuple(round(v, 1) for v in bbox)
        key = (page_num, t, bb)
        if key in seen:
            continue
        seen.add(key)
        out.append((page_num, t, bbox))
    return out


def write_excel(equations: List[EquationItem], xlsx_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Equations"

    headers = ["ID", "Page", "Equation (extracted)", "Context", "Snapshot"]
    ws.append(headers)

    # Column sizes: leave room for images
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 24

    start_row = 2
    for i, eq in enumerate(equations, start=start_row):
        ws.cell(row=i, column=1, value=eq.eq_id)
        ws.cell(row=i, column=2, value=eq.page)
        ws.cell(row=i, column=3, value=eq.text)
        ws.cell(row=i, column=4, value=eq.context)

        # Insert snapshot image
        try:
            img = XLImage(eq.image_path)
            # Fit image roughly into the cell area
            img.width = 240
            img.height = int(img.height * (240 / max(1, img.width)))
            anchor = f"E{i}"
            ws.add_image(img, anchor)
            # Increase row height to fit
            ws.row_dimensions[i].height = max(ws.row_dimensions[i].height or 15, int(img.height * 0.75))
        except Exception:
            # If image fails, continue without embedding
            pass

    # Freeze header row
    ws.freeze_panes = "A2"

    # Basic autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path.as_posix())


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Extract equation-like blocks from a PDF and export them to an Excel with embedded snapshots."
    )
    parser.add_argument(
        "--pdf",
        default=str(
            Path("Schematic-and-Layout-Design-Course-repo")
            / "HW Training - Schematic to Layout Design Course"
            / "lm5148.pdf"
        ),
        help="Path to the PDF file.",
    )
    parser.add_argument(
        "--out",
        default="lm5148_equations.xlsx",
        help="Output Excel file path.",
    )
    parser.add_argument(
        "--images-dir",
        default="lm5148_equations_images",
        help="Directory to write equation snapshot images.",
    )
    parser.add_argument(
        "--zoom",
        type=float,
        default=3.0,
        help="Render zoom factor (higher = sharper snapshots).",
    )
    parser.add_argument(
        "--min-len",
        type=int,
        default=6,
        help="Minimum extracted text length to consider as equation.",
    )
    parser.add_argument(
        "--require-equals",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="If set (default), only keep candidates containing '=', '≤', '≥', '≈', or '≠'.",
    )

    args = parser.parse_args()

    pdf_path = Path(args.pdf)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    images_dir = Path(args.images_dir)
    out_xlsx = Path(args.out)

    doc = fitz.open(pdf_path.as_posix())

    candidates: List[Tuple[int, str, Tuple[float, float, float, float], str]] = []

    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        blocks = sorted_blocks_by_position(extract_text_blocks(page))

        for idx, b in enumerate(blocks):
            txt = block_text(b)
            if not txt:
                continue
            if not looks_like_equation(txt, min_len=args.min_len, require_equals=args.require_equals):
                continue

            bbox = block_bbox(b)
            ctx = build_context(blocks, idx)
            candidates.append((page_index + 1, txt, bbox, ctx))

    # Deduplicate by (page,text,bbox)
    deduped_keyed = dedupe((p, t, bb) for (p, t, bb, _ctx) in candidates)

    # Rebuild with context (best-effort: pick first matching context)
    eq_items: List[EquationItem] = []
    eq_id = 1

    # Map back contexts
    ctx_map = {}
    for p, t, bb, ctx in candidates:
        key = (p, normalize_equation_text(t), tuple(round(v, 1) for v in bb))
        ctx_map.setdefault(key, ctx)

    for page_num, text_norm, bbox in deduped_keyed:
        page = doc.load_page(page_num - 1)
        rect = fitz.Rect(bbox)
        img_path = images_dir / f"eq_p{page_num:03d}_{eq_id:04d}.png"
        render_crop(page, rect, zoom=args.zoom, out_path=img_path)

        key = (page_num, text_norm, tuple(round(v, 1) for v in bbox))
        ctx = ctx_map.get(key, "")

        eq_items.append(
            EquationItem(
                eq_id=eq_id,
                page=page_num,
                text=text_norm,
                context=ctx,
                bbox=bbox,
                image_path=img_path.as_posix(),
            )
        )
        eq_id += 1

    write_excel(eq_items, out_xlsx)

    print(f"PDF: {pdf_path}")
    print(f"Equations found: {len(eq_items)}")
    print(f"Excel written: {out_xlsx.resolve()}")
    print(f"Images dir: {images_dir.resolve()}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
